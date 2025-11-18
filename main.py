from flask import Flask, request, jsonify, send_file, render_template_string
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy
import shutil
import logging

# 设置日志
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# 创建必要的目录
os.makedirs('uploads', exist_ok=True)
os.makedirs('outputs', exist_ok=True)


# ========== 格式复制函数 ==========

def copy_cell_style(source_cell, target_cell):
    """复制单元格样式"""
    try:
        # 复制字体
        if source_cell.font:
            target_cell.font = copy(source_cell.font)

        # 复制填充
        if source_cell.fill and source_cell.fill.fill_type:
            target_cell.fill = copy(source_cell.fill)

        # 复制边框
        if source_cell.border:
            target_cell.border = copy(source_cell.border)

        # 复制对齐
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)

        # 复制数字格式
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format

    except Exception as e:
        logger.error(f"复制单元格样式时出错: {e}")


def copy_worksheet_formatting(source_ws, target_ws):
    """复制工作表的格式设置"""
    try:
        # 复制列宽
        for col_idx in range(1, source_ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            source_dim = source_ws.column_dimensions[col_letter]
            if source_dim.width is not None:
                target_ws.column_dimensions[col_letter].width = source_dim.width

        # 复制行高
        for row_idx in range(1, min(source_ws.max_row, target_ws.max_row) + 1):
            source_dim = source_ws.row_dimensions[row_idx]
            if source_dim.height is not None:
                target_ws.row_dimensions[row_idx].height = source_dim.height

    except Exception as e:
        logger.error(f"复制工作表格式时出错: {e}")


def create_exact_copy_from_template(template_file, data_df, output_file):
    """基于模板创建精确格式副本"""
    try:
        # 复制模板文件
        shutil.copy2(template_file, output_file)
        wb = load_workbook(output_file)
        ws = wb.active

        # 清除现有数据（保留标题行）
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        # 复制标题行样式（保持加粗）
        if ws.max_row >= 1:
            for col_idx in range(1, min(len(ws[1]), len(data_df.columns)) + 1):
                source_cell = ws.cell(row=1, column=col_idx)
                target_cell = ws.cell(row=1, column=col_idx)
                copy_cell_style(source_cell, target_cell)

        # 获取数据行样式模板（使用第二行作为数据行样式模板）
        template_data_row = 2 if ws.max_row >= 2 else 1

        # 写入数据
        for row_idx, row_data in enumerate(dataframe_to_rows(data_df, index=False, header=False), 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                if col_idx > len(data_df.columns):
                    continue

                target_cell = ws.cell(row=row_idx, column=col_idx)
                target_cell.value = cell_value

                # 应用样式，但确保数据行不加粗
                if template_data_row <= ws.max_row:
                    source_cell = ws.cell(row=template_data_row, column=col_idx)
                    copy_cell_style(source_cell, target_cell)

                    # 确保数据行字体不加粗
                    if target_cell.font:
                        target_cell.font = Font(
                            name=target_cell.font.name,
                            size=target_cell.font.size,
                            bold=False,  # 数据行强制不加粗
                            italic=target_cell.font.italic,
                            underline=target_cell.font.underline,
                            color=target_cell.font.color
                        )
                else:
                    # 如果没有数据行模板，使用标题行样式但取消加粗
                    source_cell = ws.cell(row=1, column=col_idx)
                    copy_cell_style(source_cell, target_cell)
                    if target_cell.font:
                        target_cell.font = Font(
                            name=target_cell.font.name,
                            size=target_cell.font.size,
                            bold=False,  # 数据行不加粗
                            italic=target_cell.font.italic,
                            underline=target_cell.font.underline,
                            color=target_cell.font.color
                        )

        # 删除多余行
        if ws.max_row > len(data_df) + 1:
            ws.delete_rows(len(data_df) + 2, ws.max_row - len(data_df) - 1)

        wb.save(output_file)
        logger.info(f"成功创建格式化的文件: {output_file}")
        return True

    except Exception as e:
        logger.error(f"创建精确副本时出错: {str(e)}")
        # 如果模板方法失败，使用简单方法
        logger.info("尝试使用简单方法创建Excel文件")
        return create_simple_excel(data_df, output_file)


def copy_cell_style(source_cell, target_cell):
    """复制单元格样式，但允许单独控制加粗属性"""
    try:
        # 复制字体（除了加粗属性）
        if source_cell.font:
            target_cell.font = Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,  # 这个属性会在调用函数中单独控制
                italic=source_cell.font.italic,
                underline=source_cell.font.underline,
                color=source_cell.font.color
            )

        # 复制填充
        if source_cell.fill and source_cell.fill.fill_type:
            target_cell.fill = copy(source_cell.fill)

        # 复制边框
        if source_cell.border:
            target_cell.border = copy(source_cell.border)

        # 复制对齐
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)

        # 复制数字格式
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format

    except Exception as e:
        logger.error(f"复制单元格样式时出错: {e}")


def ensure_no_bold_in_data_rows(worksheet, data_df):
    """确保所有数据行字体不加粗"""
    try:
        for row in range(2, worksheet.max_row + 1):
            for col in range(1, len(data_df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.font:
                    # 创建新的字体对象，确保不加粗
                    new_font = Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=False,  # 强制不加粗
                        italic=cell.font.italic,
                        underline=cell.font.underline,
                        color=cell.font.color
                    )
                    cell.font = new_font
    except Exception as e:
        logger.error(f"确保不加粗时出错: {str(e)}")


def create_simple_excel(data_df, output_file):
    """创建简单的Excel文件（备用方案）"""
    try:
        # 直接使用pandas创建Excel文件
        data_df.to_excel(output_file, index=False, engine='openpyxl')

        # 应用基本格式
        wb = load_workbook(output_file)
        ws = wb.active

        # 设置标题行加粗
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # 确保数据行不加粗
        for row in range(2, ws.max_row + 1):
            for cell in ws[row]:
                if cell.font:
                    cell.font = Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=False,  # 数据行不加粗
                        italic=cell.font.italic,
                        underline=cell.font.underline,
                        color=cell.font.color
                    )

        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_file)
        return True
    except Exception as e:
        logger.error(f"创建简单Excel时出错: {str(e)}")
        return False

# ========== 其他工具函数 ==========

def get_colleges_from_data(df):
    """从数据中获取所有学院列表"""
    college_columns = [col for col in df.columns if '院系' in col or '学院' in col or 'Address' in col]
    if college_columns:
        college_column = college_columns[0]
    else:
        college_column = df.columns[1] if len(df.columns) > 1 else df.columns[0]
    colleges = df[college_column].dropna().unique()
    return colleges.tolist(), college_column


def reset_serial_numbers(data_df):
    """重置序号列"""
    number_columns = [col for col in data_df.columns if any(keyword in str(col) for keyword in
                                                            ['Number', '序号', '编号', 'No.', 'NO', '编号'])]

    if number_columns:
        number_column = number_columns[0]
        data_df[number_column] = range(1, len(data_df) + 1)
    else:
        first_col_name = str(data_df.columns[0])
        if any(keyword in first_col_name for keyword in ['Number', '序号', '编号', 'No.', 'NO']):
            data_df.iloc[:, 0] = range(1, len(data_df) + 1)

    return data_df


def get_safe_filename(name):
    """生成安全的文件名"""
    unsafe_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
    safe_name = name
    for char in unsafe_chars:
        safe_name = safe_name.replace(char, '_')
    if len(safe_name) > 100:
        safe_name = safe_name[:100]
    return safe_name


def get_unique_filename(directory, base_name, extension):
    """生成唯一的文件名"""
    safe_base_name = get_safe_filename(base_name)
    counter = 1
    file_path = os.path.join(directory, f"{safe_base_name}{extension}")

    while os.path.exists(file_path):
        file_path = os.path.join(directory, f"{safe_base_name}_{counter}{extension}")
        counter += 1

    return file_path


# ========== 核心功能函数 ==========

def filter_by_college_only(main_file_path, selected_college, college_column):
    """仅按学院筛选（不进行查重）"""
    try:
        logger.info(f"开始学院筛选: {selected_college}")
        main_df = pd.read_excel(main_file_path)
        original_count = len(main_df)

        logger.info(f"原始数据记录数: {original_count}")
        logger.info(f"学院列: {college_column}")

        # 筛选指定学院
        college_papers = main_df[main_df[college_column] == selected_college].copy()
        remaining_papers = main_df[main_df[college_column] != selected_college].copy()

        logger.info(f"筛选出的学院论文数: {len(college_papers)}")
        logger.info(f"剩余论文数: {len(remaining_papers)}")

        # 重置序号
        college_papers = reset_serial_numbers(college_papers)
        remaining_papers = reset_serial_numbers(remaining_papers)

        return college_papers, remaining_papers, original_count, None

    except Exception as e:
        logger.error(f"学院筛选时出错: {str(e)}")
        return None, None, None, f"学院筛选时出错: {str(e)}"


def correct_deduplicate_and_filter(check_file_path, main_file_path, selected_college, college_column):
    """修正的查重逻辑"""
    try:
        logger.info("=== 开始查重处理 ===")

        # 读取文件
        check_df = pd.read_excel(check_file_path)
        main_df = pd.read_excel(main_file_path)

        logger.info(f"查重文件记录数: {len(check_df)}")
        logger.info(f"主文件记录数: {len(main_df)}")

        # 检查WOS列
        if 'WOS Accession Number' not in check_df.columns:
            return None, None, None, None, "查重文件中找不到'WOS Accession Number'列"
        if 'WOS Accession Number' not in main_df.columns:
            return None, None, None, None, "主文件中找不到'WOS Accession Number'列"

        # 获取主文件中的WOS编号
        main_wos_numbers = set(main_df['WOS Accession Number'].dropna().unique())

        # 从查重文件中删除重复数据
        deduplicated_df = check_df[~check_df['WOS Accession Number'].isin(main_wos_numbers)]

        logger.info(f"去重后记录数: {len(deduplicated_df)}")
        logger.info(f"删除的记录数: {len(check_df) - len(deduplicated_df)}")

        # 筛选指定学院
        if college_column not in deduplicated_df.columns:
            return None, None, None, None, f"查重文件中找不到学院列: {college_column}"

        college_papers = deduplicated_df[deduplicated_df[college_column] == selected_college].copy()
        remaining_papers = deduplicated_df[deduplicated_df[college_column] != selected_college].copy()

        logger.info(f"学院'{selected_college}'论文数: {len(college_papers)}")
        logger.info(f"剩余论文数: {len(remaining_papers)}")

        # 重置序号
        college_papers = reset_serial_numbers(college_papers)
        remaining_papers = reset_serial_numbers(remaining_papers)

        return college_papers, remaining_papers, len(check_df), len(check_df) - len(deduplicated_df), None

    except Exception as e:
        logger.error(f"查重筛选错误: {str(e)}")
        return None, None, None, None, f"查重筛选时出错: {str(e)}"


def get_correct_deduplicated_stats(check_file_path, main_file_path, college_column):
    """获取正确的查重后学院统计"""
    try:
        logger.info("=== 获取查重统计 ===")

        # 读取文件
        check_df = pd.read_excel(check_file_path)
        main_df = pd.read_excel(main_file_path)

        logger.info(f"统计 - 查重文件: {len(check_df)} 条")
        logger.info(f"统计 - 主文件: {len(main_df)} 条")

        # 检查WOS列
        if 'WOS Accession Number' not in check_df.columns or 'WOS Accession Number' not in main_df.columns:
            logger.error("错误: 缺少WOS列")
            return {}

        # 获取主文件中的WOS编号
        main_wos_numbers = set(main_df['WOS Accession Number'].dropna().unique())

        # 从查重文件中删除重复数据
        deduplicated_df = check_df[~check_df['WOS Accession Number'].isin(main_wos_numbers)]

        logger.info(f"统计 - 去重后: {len(deduplicated_df)} 条")
        logger.info(f"统计 - 删除数: {len(check_df) - len(deduplicated_df)} 条")

        # 获取学院统计
        if college_column not in deduplicated_df.columns:
            logger.error(f"错误: 找不到学院列 {college_column}")
            return {}

        college_counts = deduplicated_df[college_column].value_counts().to_dict()
        logger.info(f"学院统计: {college_counts}")

        return college_counts

    except Exception as e:
        logger.error(f"获取查重统计错误: {str(e)}")
        return {}


# ========== Flask 路由 ==========

@app.route('/')
def index():
    """主页面"""
    html_content = '''
    <!DOCTYPE html>
    <html lang="zh-CN">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>学院论文连续筛选平台</title>
        <style>
            * { box-sizing: border-box; margin: 0; padding: 0; }
            body {
                font-family: 'Microsoft YaHei', Arial, sans-serif;
                line-height: 1.6;
                color: #333;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                min-height: 100vh;
                padding: 20px;
            }
            .container {
                max-width: 1200px;
                margin: 0 auto;
                background: white;
                border-radius: 15px;
                box-shadow: 0 10px 30px rgba(0,0,0,0.2);
                overflow: hidden;
            }
            .header {
                background: linear-gradient(135deg, #2c3e50, #3498db);
                color: white;
                padding: 30px;
                text-align: center;
            }
            .header h1 {
                font-size: 2.5em;
                margin-bottom: 10px;
            }
            .content-area {
                padding: 30px;
            }

            /* 上传区域样式 */
            .upload-section {
                display: grid;
                grid-template-columns: 1fr;
                gap: 20px;
                margin: 20px 0;
            }

            .upload-area {
                border: 3px dashed #3498db;
                border-radius: 15px;
                padding: 30px;
                text-align: center;
                background: #f8f9fa;
                transition: all 0.3s ease;
                cursor: pointer;
                min-height: 150px;
                display: flex;
                flex-direction: column;
                justify-content: center;
                align-items: center;
                position: relative;
            }

            .upload-area:hover {
                border-color: #2980b9;
                background: #e8f4fc;
                transform: translateY(-2px);
            }

            .upload-icon { 
                font-size: 36px; 
                margin-bottom: 10px;
            }

            .file-info {
                background: #e8f4fc;
                padding: 12px;
                border-radius: 8px;
                margin: 10px 0;
                width: 100%;
                text-align: left;
            }

            /* 选项区域 */
            .option-section {
                background: #f8f9fa;
                padding: 25px;
                border-radius: 12px;
                margin: 25px 0;
                text-align: center;
            }

            .option-toggle {
                display: flex;
                justify-content: center;
                align-items: center;
                gap: 20px;
                margin: 20px 0;
            }

            .toggle-switch {
                position: relative;
                display: inline-block;
                width: 70px;
                height: 38px;
            }

            .toggle-switch input {
                opacity: 0;
                width: 0;
                height: 0;
            }

            .slider {
                position: absolute;
                cursor: pointer;
                top: 0;
                left: 0;
                right: 0;
                bottom: 0;
                background-color: #ccc;
                transition: .4s;
                border-radius: 34px;
            }

            .slider:before {
                position: absolute;
                content: "";
                height: 30px;
                width: 30px;
                left: 4px;
                bottom: 4px;
                background-color: white;
                transition: .4s;
                border-radius: 50%;
            }

            input:checked + .slider {
                background: linear-gradient(135deg, #3498db, #2980b9);
            }

            input:checked + .slider:before {
                transform: translateX(32px);
            }

            /* 按钮样式 */
            .button {
                background: linear-gradient(135deg, #3498db, #2980b9);
                color: white;
                border: none;
                padding: 12px 25px;
                border-radius: 25px;
                cursor: pointer;
                font-size: 16px;
                transition: all 0.3s ease;
                margin: 8px;
                font-weight: 600;
            }

            .button:hover {
                transform: translateY(-2px);
            }

            .button:disabled {
                background: #bdc3c7;
                cursor: not-allowed;
                transform: none;
            }

            .button.process {
                background: linear-gradient(135deg, #27ae60, #229954);
                font-size: 17px;
                padding: 15px 35px;
            }

            .button.download {
                background: linear-gradient(135deg, #e74c3c, #c0392b);
            }

            .button.continue {
                background: linear-gradient(135deg, #f39c12, #e67e22);
            }

            .hidden { display: none; }

            /* 学院列表样式 */
            .college-list {
                display: grid;
                grid-template-columns: repeat(auto-fill, minmax(200px, 1fr));
                gap: 12px;
                margin: 20px 0;
            }

            .college-item {
                padding: 15px;
                border: 2px solid #e0e0e0;
                border-radius: 10px;
                cursor: pointer;
                text-align: center;
                transition: all 0.3s ease;
                background: white;
            }

            .college-item:hover {
                border-color: #3498db;
                transform: translateY(-2px);
            }

            .college-item.selected {
                border-color: #27ae60;
                background: #27ae60;
                color: white;
            }

            .current-file-info {
                background: #e8f6ff;
                padding: 20px;
                border-radius: 10px;
                margin: 15px 0;
            }

            .message {
                padding: 12px;
                border-radius: 8px;
                margin: 12px 0;
                text-align: center;
                font-weight: bold;
            }

            .loading { background: #d4edfc; color: #004085; }
            .error { background: #f8d7da; color: #721c24; }
            .success { background: #d4edda; color: #155724; }

            .result-section {
                text-align: center;
                margin: 25px 0;
            }

            .stats {
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
                gap: 15px;
                margin: 25px 0;
            }

            .stat-item {
                background: white;
                padding: 20px;
                border-radius: 10px;
                text-align: center;
                box-shadow: 0 5px 15px rgba(0,0,0,0.1);
                border: 2px solid #3498db;
            }

            .stat-number {
                font-size: 2.2em;
                font-weight: bold;
                color: #3498db;
                margin-bottom: 8px;
            }

            input[type="file"] {
                position: absolute;
                width: 100%;
                height: 100%;
                top: 0;
                left: 0;
                opacity: 0;
                cursor: pointer;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>🎓 学院论文连续筛选平台</h1>
                <p>可选择是否进行查重，按学院连续筛选</p>
            </div>

            <div class="content-area">
                <!-- 查重选项 -->
                <div class="option-section">
                    <h3>🔧 处理选项</h3>
                    <div class="option-toggle">
                        <span>仅筛选</span>
                        <label class="toggle-switch">
                            <input type="checkbox" id="deduplicationToggle">
                            <span class="slider"></span>
                        </label>
                        <span>查重后筛选</span>
                    </div>
                    <p id="optionDescription">当前模式：仅按学院筛选（不进行查重）</p>
                </div>

                <!-- 文件上传区域 -->
                <div class="upload-section">
                    <!-- 主文件上传区域 -->
                    <div id="mainFileUploadSection">
                        <div class="upload-area" onclick="document.getElementById('mainFileInput').click()">
                            <div class="upload-icon">📁</div>
                            <h3>主表格文件</h3>
                            <p>需要进行筛选的主要文件</p>
                            <input type="file" id="mainFileInput" accept=".xlsx,.xls" class="hidden">
                            <div id="mainFileInfo" class="file-info hidden"></div>
                        </div>
                    </div>

                    <!-- 查重文件上传区域（默认隐藏） -->
                    <div id="checkFileUploadSection" class="hidden">
                        <div class="upload-area" onclick="document.getElementById('checkFileInput').click()">
                            <div class="upload-icon">📊</div>
                            <h3>查重表格文件</h3>
                            <p>完整数据集 - 作为筛选基础</p>
                            <input type="file" id="checkFileInput" accept=".xlsx,.xls" class="hidden">
                            <div id="checkFileInfo" class="file-info hidden"></div>
                        </div>
                    </div>
                </div>

                <!-- 处理区域 -->
                <div id="processingSection" class="hidden">
                    <div id="currentFileInfo" class="current-file-info"></div>

                    <!-- 学院选择 -->
                    <div id="collegeSelectionSection">
                        <h3>选择要筛选出的学院：</h3>
                        <div id="collegeList" class="college-list"></div>
                        <div style="text-align: center; margin-top: 20px;">
                            <button id="processCollegeBtn" class="button process" disabled>开始筛选</button>
                        </div>
                    </div>

                    <!-- 结果展示 -->
                    <div id="resultSection" class="result-section hidden">
                        <h2>✅ 筛选完成！</h2>
                        <div id="resultStats" class="stats"></div>
                        <div style="text-align: center; margin-top: 25px;">
                            <button id="downloadCollegeBtn" class="button download">下载筛选出的论文表格</button>
                            <button id="downloadRemainingBtn" class="button download">下载剩余论文表格</button>
                            <button id="continueFilterBtn" class="button continue">继续筛选剩余数据</button>
                        </div>
                    </div>
                </div>

                <!-- 消息区域 -->
                <div id="messageArea"></div>
            </div>
        </div>

        <script>
            // 全局变量
            let currentFiles = {
                mainFile: null,
                checkFile: null
            };
            let currentResult = null;
            let selectedCollege = null;
            let collegeStatistics = {};
            let useDeduplication = false;
            let selectedCollegesHistory = [];
            let cumulativeCollegeStats = {};

            // 初始化
            document.addEventListener('DOMContentLoaded', function() {
                // 绑定事件
                document.getElementById('deduplicationToggle').addEventListener('change', toggleDeduplication);
                document.getElementById('mainFileInput').addEventListener('change', handleMainFileUpload);
                document.getElementById('checkFileInput').addEventListener('change', handleCheckFileUpload);
                document.getElementById('processCollegeBtn').addEventListener('click', processCollegeData);
                document.getElementById('downloadCollegeBtn').addEventListener('click', downloadCollegeFile);
                document.getElementById('downloadRemainingBtn').addEventListener('click', downloadRemainingFile);
                document.getElementById('continueFilterBtn').addEventListener('click', continueFiltering);
            });

            // 切换查重模式
            function toggleDeduplication() {
                useDeduplication = document.getElementById('deduplicationToggle').checked;
                const checkFileSection = document.getElementById('checkFileUploadSection');
                const optionDescription = document.getElementById('optionDescription');

                if (useDeduplication) {
                    checkFileSection.classList.remove('hidden');
                    optionDescription.textContent = '当前模式：查重后筛选';
                } else {
                    checkFileSection.classList.add('hidden');
                    optionDescription.textContent = '当前模式：仅按学院筛选';
                }

                // 重置状态
                currentFiles.checkFile = null;
                document.getElementById('checkFileInfo').classList.add('hidden');
                document.getElementById('checkFileInfo').innerHTML = '';

                if (currentFiles.mainFile) {
                    checkProcessingReadyState();
                }
            }

            // 处理主文件上传
            function handleMainFileUpload(event) {
                handleFileUpload(event, 'mainFile');
            }

            // 处理查重文件上传
            function handleCheckFileUpload(event) {
                handleFileUpload(event, 'checkFile');
            }

            // 处理文件上传
            function handleFileUpload(event, fileType) {
                const file = event.target.files[0];
                if (!file) return;

                showMessage('正在验证文件...', 'loading');

                const formData = new FormData();
                formData.append('file', file);
                formData.append('file_type', fileType);

                fetch('/upload', {
                    method: 'POST',
                    body: formData
                })
                .then(response => response.json())
                .then(result => {
                    if (result.success) {
                        currentFiles[fileType] = result;
                        displayFileInfo(fileType, result);

                        if (fileType === 'mainFile') {
                            document.getElementById('mainFileUploadSection').classList.add('hidden');
                            updateCurrentFileInfo();
                            cumulativeCollegeStats = {};
                        }

                        checkProcessingReadyState();
                        showMessage(`${fileType === 'mainFile' ? '主' : '查重'}文件上传成功！`, 'success');
                    } else {
                        showMessage(result.error, 'error');
                        currentFiles[fileType] = null;
                        document.getElementById(fileType + 'Input').value = '';
                    }
                })
                .catch(error => {
                    showMessage('上传文件时出错: ' + error.message, 'error');
                    currentFiles[fileType] = null;
                    document.getElementById(fileType + 'Input').value = '';
                });
            }

            // 显示文件信息
            function displayFileInfo(fileType, info) {
                const infoDiv = document.getElementById(fileType + 'Info');
                const fileLabel = fileType === 'mainFile' ? '主文件' : '查重文件';

                let html = `
                    <h4>${fileLabel}信息：</h4>
                    <p><strong>文件名：</strong> ${info.filename}</p>
                    <p><strong>记录数：</strong> ${info.record_count} 条</p>
                `;

                if (fileType === 'mainFile') {
                    html += `<p><strong>学院数：</strong> ${info.colleges ? info.colleges.length : 0} 个</p>`;
                }

                infoDiv.innerHTML = html;
                infoDiv.classList.remove('hidden');
            }

            // 检查是否可以开始处理
            function checkProcessingReadyState() {
                const mainReady = currentFiles.mainFile !== null;
                const checkReady = !useDeduplication || (useDeduplication && currentFiles.checkFile !== null);

                if (mainReady && checkReady) {
                    showProcessingSection();
                    getCollegeStatistics();
                }
            }

            // 显示处理区域
            function showProcessingSection() {
                document.getElementById('processingSection').classList.remove('hidden');
                document.getElementById('collegeSelectionSection').classList.remove('hidden');
                updateCurrentFileInfo();
            }

            // 更新当前文件信息
            function updateCurrentFileInfo() {
                const infoDiv = document.getElementById('currentFileInfo');
                let html = `
                    <h4>当前处理状态：</h4>
                    <p><strong>处理模式：</strong> ${useDeduplication ? '查重后筛选' : '仅筛选'}</p>
                `;

                if (useDeduplication) {
                    html += `
                        <p><strong>查重文件：</strong> ${currentFiles.checkFile.filename} (${currentFiles.checkFile.record_count} 条)</p>
                        <p><strong>主文件：</strong> ${currentFiles.mainFile.filename} (${currentFiles.mainFile.record_count} 条)</p>
                    `;
                } else {
                    html += `<p><strong>主文件：</strong> ${currentFiles.mainFile.filename} (${currentFiles.mainFile.record_count} 条)</p>`;
                }

                infoDiv.innerHTML = html;
            }

            // 获取学院统计
            function getCollegeStatistics() {
                if (!currentFiles.mainFile) return;

                showMessage('正在统计各学院论文数量...', 'loading');

                const requestData = {
                    main_file_path: currentFiles.mainFile.file_path,
                    college_column: currentFiles.mainFile.college_column,
                    use_deduplication: useDeduplication
                };

                if (useDeduplication && currentFiles.checkFile) {
                    requestData.check_file_path = currentFiles.checkFile.file_path;
                }

                fetch('/get-college-statistics', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify(requestData)
                })
                .then(response => response.json())
                .then(result => {
                    if (result.success) {
                        collegeStatistics = result.college_stats;
                        displayCollegeList(result.college_stats);
                        showMessage('学院统计完成！', 'success');
                    } else {
                        showMessage(result.error, 'error');
                    collegeStatistics = {};
                    displayCollegeList({});
                    document.getElementById('collegeSelectionSection').classList.add('hidden');
                    document.getElementById('resultSection').classList.add('hidden');
                    document.getElementById('processCollegeBtn').disabled = true;
                    document.getElementById('processCollegeBtn').textContent = '开始筛选';
                    selectedCollege = null;
                }
                })
                .catch(error => {
                    showMessage('获取学院统计时出错: ' + error.message, 'error');
                    collegeStatistics = {};
                    displayCollegeList({});
                    document.getElementById('collegeSelectionSection').classList.add('hidden');
                    document.getElementById('resultSection').classList.add('hidden');
                    document.getElementById('processCollegeBtn').disabled = true;
                    document.getElementById('processCollegeBtn').textContent = '开始筛选';
                    selectedCollege = null;
                });
            }

            // 显示学院选择列表
            function displayCollegeList(stats) {
                const collegeList = document.getElementById('collegeList');
                collegeList.innerHTML = '';

                if (Object.keys(stats).length === 0) {
                    collegeList.innerHTML = '<p style="text-align: center; color: #666; grid-column: 1 / -1;">没有找到可筛选的学院数据</p>';
                    document.getElementById('collegeSelectionSection').classList.add('hidden');
                    return;
                }

                for (const [college, count] of Object.entries(stats)) {
                    if (count > 0) {
                        const collegeItem = document.createElement('div');
                        collegeItem.className = 'college-item';
                        const cumulativeCount = cumulativeCollegeStats[college] || 0;

                        collegeItem.innerHTML = `
                            <div style="font-size: 1.1em; font-weight: bold; margin-bottom: 5px;">${college}</div>
                            <div style="font-size: 0.85em; color: #666;">
                                <div>当前剩余: <strong style="color: #27ae60">${count}</strong> 篇</div>
                                ${cumulativeCount > 0 ? `<div style="color: #e67e22; margin-top: 5px;">累计: ${cumulativeCount} 篇</div>` : ''}
                            </div>
                        `;
                        collegeItem.onclick = () => selectCollege(college, count);
                        collegeList.appendChild(collegeItem);
                    }
                }

                document.getElementById('collegeSelectionSection').classList.remove('hidden');
            }

            // 选择学院
            function selectCollege(college, count) {
                selectedCollege = college;
                document.querySelectorAll('.college-item').forEach(item => {
                    item.classList.remove('selected');
                    if (item.textContent.includes(college)) {
                        item.classList.add('selected');
                    }
                });

                const processBtn = document.getElementById('processCollegeBtn');
                processBtn.disabled = false;
                const cumulativeCount = cumulativeCollegeStats[college] || 0;

                if (cumulativeCount > 0) {
                    processBtn.textContent = `筛选 ${college} 的数据 (${count}篇, 累计: ${cumulativeCount}篇)`;
                } else {
                    processBtn.textContent = `筛选 ${college} 的数据 (${count}篇)`;
                }
            }

            // 处理学院数据
            function processCollegeData() {
                if (!selectedCollege || !currentFiles.mainFile) return;

                showMessage('正在筛选数据，请稍候...', 'loading');

                const requestData = {
                    main_file_path: currentFiles.mainFile.file_path,
                    selected_college: selectedCollege,
                    college_column: currentFiles.mainFile.college_column,
                    use_deduplication: useDeduplication
                };

                if (useDeduplication && currentFiles.checkFile) {
                    requestData.check_file_path = currentFiles.checkFile.file_path;
                }

                fetch('/process-college', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify(requestData)
                })
                .then(response => response.json())
                .then(result => {
                    if (result.success) {
                        currentResult = result;
                        selectedCollegesHistory.push(selectedCollege);

                        // 更新累计统计
                        cumulativeCollegeStats[selectedCollege] = (cumulativeCollegeStats[selectedCollege] || 0) + result.college_count;

                        showResults(result);
                        showMessage('筛选完成！', 'success');
                    } else {
                        showMessage(result.error, 'error');
                    }
                })
                .catch(error => {
                    showMessage('处理数据时出错: ' + error.message, 'error');
                });
            }

            // 显示结果
            function showResults(result) {
                const statsDiv = document.getElementById('resultStats');
                const cumulativeCount = cumulativeCollegeStats[selectedCollege] || 0;

                let statsHTML = `
                    <div class="stat-item">
                        <div class="stat-number">${result.original_count}</div>
                        <div>${useDeduplication ? '查重文件' : '原始'}论文数</div>
                    </div>
                `;

                if (useDeduplication) {
                    statsHTML += `
                        <div class="stat-item">
                            <div class="stat-number">${result.removed_count}</div>
                            <div>删除重复数</div>
                        </div>
                        <div class="stat-item">
                            <div class="stat-number">${result.original_count - result.removed_count}</div>
                            <div>去重后总数</div>
                        </div>
                    `;
                }

                statsHTML += `
                    <div class="stat-item">
                        <div class="stat-number" style="color: #27ae60">${result.college_count}</div>
                        <div>${selectedCollege}论文数</div>
                    </div>
                    <div class="stat-item">
                        <div class="stat-number">${result.remaining_count}</div>
                        <div>剩余论文数</div>
                    </div>
                    <div class="stat-item">
                        <div class="stat-number" style="color: #e67e22">${cumulativeCount}</div>
                        <div>${selectedCollege}累计总数</div>
                    </div>
                `;

                statsDiv.innerHTML = statsHTML;
                document.getElementById('resultSection').classList.remove('hidden');
            }

            // 下载文件
            function downloadCollegeFile() {
                if (!currentResult || !currentResult.college_file) {
                    showMessage('没有可下载的文件', 'error');
                    return;
                }
                downloadFile(currentResult.college_file, '学院文件');
            }

            function downloadRemainingFile() {
                if (!currentResult || !currentResult.remaining_file) {
                    showMessage('没有可下载的文件', 'error');
                    return;
                }
                downloadFile(currentResult.remaining_file, '剩余文件');
            }

            function downloadFile(filename, fileType) {
                const link = document.createElement('a');
                link.href = `/download/${filename}`;
                link.download = filename;
                link.style.display = 'none';
                document.body.appendChild(link);
                link.click();
                document.body.removeChild(link);
                showMessage(`开始下载${fileType}...`, 'success');
            }

            // 继续筛选
            function continueFiltering() {
                if (!currentResult) return;

                document.getElementById('resultSection').classList.add('hidden');
                selectedCollege = null;
                document.getElementById('processCollegeBtn').disabled = true;
                document.getElementById('processCollegeBtn').textContent = '开始筛选';

                updateCurrentFileInfo();
                getCollegeStatistics();
                showMessage('可以继续筛选下一个学院', 'success');
            }

            // 显示消息
            function showMessage(text, type) {
                const messageArea = document.getElementById('messageArea');
                messageArea.innerHTML = `<div class="message ${type}">${text}</div>`;

                if (type === 'success') {
                    setTimeout(() => {
                        messageArea.innerHTML = '';
                    }, 5000);
                }
            }
        </script>
    </body>
    </html>
    '''
    return render_template_string(html_content)


@app.route('/upload', methods=['POST'])
def upload_file():
    """上传Excel文件并验证"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '没有选择文件'})

    file = request.files['file']
    file_type = request.form.get('file_type', 'mainFile')

    if file.filename == '':
        return jsonify({'success': False, 'error': '没有选择文件'})

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'error': '请上传Excel文件'})

    try:
        # 保存文件
        os.makedirs('uploads', exist_ok=True)
        file_path = os.path.join('uploads', file.filename)
        file.save(file_path)

        # 读取Excel文件
        df = pd.read_excel(file_path)

        # 获取学院信息
        colleges, college_column = get_colleges_from_data(df)

        response_data = {
            'success': True,
            'filename': file.filename,
            'file_path': file_path,
            'record_count': len(df),
            'colleges': colleges,
            'college_column': college_column,
            'has_wos': 'WOS Accession Number' in df.columns
        }

        return jsonify(response_data)

    except Exception as e:
        logger.error(f"上传文件时出错: {str(e)}")
        return jsonify({'success': False, 'error': f'处理文件时出错: {str(e)}'})


@app.route('/get-college-statistics', methods=['POST'])
def get_college_stats():
    """获取学院统计信息"""
    data = request.json
    main_file_path = data.get('main_file_path')
    college_column = data.get('college_column')
    use_deduplication = data.get('use_deduplication', False)
    check_file_path = data.get('check_file_path')

    try:
        if use_deduplication and check_file_path:
            college_stats = get_correct_deduplicated_stats(check_file_path, main_file_path, college_column)
        else:
            df = pd.read_excel(main_file_path)
            college_stats = df[college_column].value_counts().to_dict()

        return jsonify({
            'success': True,
            'college_stats': college_stats
        })

    except Exception as e:
        logger.error(f"获取学院统计时出错: {str(e)}")
        return jsonify({'success': False, 'error': f'获取学院统计时出错: {str(e)}'})


@app.route('/process-college', methods=['POST'])
def process_college_data():
    """处理学院数据筛选"""
    data = request.json
    main_file_path = data.get('main_file_path')
    selected_college = data.get('selected_college')
    college_column = data.get('college_column')
    use_deduplication = data.get('use_deduplication', False)
    check_file_path = data.get('check_file_path')

    try:
        logger.info(f"开始处理学院数据: {selected_college}")

        if use_deduplication and check_file_path:
            logger.info("使用查重模式")
            result = correct_deduplicate_and_filter(check_file_path, main_file_path, selected_college, college_column)
            if result[4] is not None:  # 错误信息
                return jsonify({'success': False, 'error': result[4]})
            college_papers, remaining_papers, original_count, removed_count, _ = result
            template_file = check_file_path  # 使用查重文件作为模板
        else:
            logger.info("使用普通筛选模式")
            result = filter_by_college_only(main_file_path, selected_college, college_column)
            college_papers, remaining_papers, original_count, error_msg = result
            if error_msg:
                return jsonify({'success': False, 'error': error_msg})
            removed_count = 0
            template_file = main_file_path  # 使用主文件作为模板

        if college_papers is None or len(college_papers) == 0:
            return jsonify({'success': False, 'error': f'未找到属于"{selected_college}"的论文'})

        # 生成输出文件
        safe_college_name = get_safe_filename(selected_college)
        college_file = get_unique_filename('outputs', safe_college_name, ".xlsx")
        remaining_file = get_unique_filename('outputs', "剩余数据", ".xlsx")

        logger.info(f"创建学院文件: {college_file}")
        logger.info(f"创建剩余文件: {remaining_file}")

        # 使用模板创建格式化的Excel文件
        success1 = create_exact_copy_from_template(template_file, college_papers, college_file)
        success2 = create_exact_copy_from_template(template_file, remaining_papers, remaining_file)

        if success1 and success2:
            response_data = {
                'success': True,
                'college_file': os.path.basename(college_file),
                'remaining_file': os.path.basename(remaining_file),
                'college_count': len(college_papers),
                'remaining_count': len(remaining_papers),
                'original_count': original_count,
                'removed_count': removed_count
            }
            logger.info(f"处理成功: {response_data}")
            return jsonify(response_data)
        else:
            logger.error("Excel文件创建失败")
            return jsonify({'success': False, 'error': '处理文件时出错'})

    except Exception as e:
        logger.error(f"处理数据时出错: {str(e)}")
        return jsonify({'success': False, 'error': f'处理数据时出错: {str(e)}'})


@app.route('/download/<filename>')
def download_file(filename):
    """下载处理后的文件"""
    try:
        file_path = os.path.join('outputs', filename)

        if not os.path.exists(file_path):
            return jsonify({'success': False, 'error': f'文件不存在: {filename}'}), 404

        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f"下载文件时出错: {str(e)}")
        return jsonify({'success': False, 'error': f'下载文件时出错: {str(e)}'}), 500


if __name__ == '__main__':
    # 确保输出目录存在
    os.makedirs('outputs', exist_ok=True)
    os.makedirs('uploads', exist_ok=True)

    logger.info("启动Flask应用...")
    app.run(debug=True, host='0.0.0.0', port=5011)
